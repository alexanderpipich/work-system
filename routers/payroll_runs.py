from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy.orm import Session

from access import apply_shift_scope, get_economist_cities
from audit_helpers import create_audit_log
from document_helpers import citizenship_display
from dependencies import current_user, get_db
from models import PayrollRun, PayrollRunItem, Requisite, Shift, User
from payroll_closing import close_payroll_run as close_payroll_run_financials
from rbac import canonical_role, has_permission, is_superadmin, require_permission
from schedule_grid import days_between
from time_helpers import now_utc
from utils import normalize_text


router = APIRouter()
templates = Jinja2Templates(directory="templates")
require_payroll_delete = require_permission("payroll.delete", audit_denied=True)
require_payroll_close = require_permission("payroll.close", audit_denied=True)
require_payroll_view = require_permission("payroll.view", audit_denied=True)
require_payroll_export = require_permission("payroll.export", audit_denied=True)
_payroll_manage = require_permission("payroll.manage")


def _item_key(employee, store, service, shift_date):
    return (
        normalize_text(employee),
        normalize_text(store),
        normalize_text(service),
        shift_date,
    )


def _allowed_item_keys(session, allowed_cities):
    if not allowed_cities:
        return set()

    rows = session.query(
        Shift.employee,
        Shift.store,
        Shift.service,
        Shift.shift_date,
    ).filter(Shift.city.in_(allowed_cities)).distinct().all()

    return {_item_key(row[0], row[1], row[2], row[3]) for row in rows}


def _filter_items_by_cities(session, items, allowed_cities):
    allowed_keys = _allowed_item_keys(session, allowed_cities)
    return [
        item
        for item in items
        if _item_key(
            item.employee_name,
            item.store,
            item.service,
            item.shift_date,
        ) in allowed_keys
    ]


def _all_items_accessible_to_cities(session, items, allowed_cities):
    if not items or not allowed_cities:
        return False
    return len(_filter_items_by_cities(session, items, allowed_cities)) == len(items)

def build_run_rows(session, allowed_cities=None):
    runs = session.query(PayrollRun).order_by(PayrollRun.id.desc()).all()
    run_rows = []

    for run in runs:
        items = session.query(PayrollRunItem).filter(
            PayrollRunItem.run_id == run.id
        ).all()
        if allowed_cities is not None:
            items = _filter_items_by_cities(session, items, allowed_cities)
            if not items:
                continue

        total_hours = sum(item.hours or 0 for item in items)
        total_amount = sum(item.total_amount or 0 for item in items)

        creator = None
        if run.created_by:
            creator = session.query(User).filter(User.id == run.created_by).first()

        run_rows.append({
            "run": run,
            "items_count": len(items),
            "total_hours": total_hours,
            "total_amount": total_amount,
            "creator": creator.employee_name if creator else "—"
        })

    return run_rows


def get_run_items(session, run_id, allowed_cities=None):
    run = session.query(PayrollRun).filter(PayrollRun.id == run_id).first()

    if not run:
        return None, []

    items = session.query(PayrollRunItem).filter(
        PayrollRunItem.run_id == run.id
    ).order_by(
        PayrollRunItem.employee_name.asc(),
        PayrollRunItem.store.asc(),
        PayrollRunItem.service.asc(),
        PayrollRunItem.shift_date.asc()
    ).all()

    if allowed_cities is not None:
        items = _filter_items_by_cities(session, items, allowed_cities)

    return run, items



def _allowed_item_keys_for_user(session, user):
    query = session.query(Shift.employee, Shift.store, Shift.service, Shift.shift_date).distinct()
    rows = apply_shift_scope(query, session, user, Shift).all()
    return {_item_key(row[0], row[1], row[2], row[3]) for row in rows}


def _filter_items_for_user(session, items, user):
    allowed_keys = _allowed_item_keys_for_user(session, user)
    return [item for item in items if _item_key(item.employee_name, item.store, item.service, item.shift_date) in allowed_keys]


def build_run_rows_for_user(session, user):
    rows = []
    for run in session.query(PayrollRun).order_by(PayrollRun.id.desc()).all():
        items = _filter_items_for_user(session, session.query(PayrollRunItem).filter(PayrollRunItem.run_id == run.id).all(), user)
        if not items:
            continue
        creator = session.query(User).filter(User.id == run.created_by).first() if run.created_by else None
        rows.append({"run": run, "items_count": len(items), "total_hours": sum(item.hours or 0 for item in items), "total_amount": sum(item.total_amount or 0 for item in items), "creator": creator.employee_name if creator else "—"})
    return rows


def get_run_items_for_user(session, run_id, user):
    run, items = get_run_items(session, run_id)
    return run, _filter_items_for_user(session, items, user)


def build_run_day_rows(run, items):
    """Строки зафиксированного табеля в виде «по дням» (вид_табелей).

    Агрегирует ИСТОРИЧЕСКИЕ строки прогона (PayrollRunItem) по (ФИО, магазин,
    услуга): часы раскладываются по дням периода прогона, суммы берутся из
    прогона как есть — НИЧЕГО не пересчитываем текущими ставками. Колонка
    «корректировки/ЛМК» = total_amount − amount (авто+ручные−ЛМК, зафиксированные).
    Общий ИТОГ (часы/суммы) равен sum по items — не меняется."""
    days = days_between(run.date_from, run.date_to) if run else []
    day_set = set(days)
    table = {}
    order = []

    for item in items:
        key = (item.employee_name, item.store, item.service)
        row = table.get(key)
        if row is None:
            row = {
                "employee_name": item.employee_name,
                "store": item.store,
                "service": item.service,
                "days": {day: 0 for day in days},
                "hours": 0,
                "rate": item.rate or 0,
                "amount": 0,
                "total_amount": 0,
                "corrections": 0,
            }
            table[key] = row
            order.append(key)

        if item.shift_date in day_set:
            row["days"][item.shift_date] += item.hours or 0
        row["hours"] += item.hours or 0
        if (item.rate or 0) and not row["rate"]:
            row["rate"] = item.rate or 0
        row["amount"] += item.amount or 0
        row["total_amount"] += item.total_amount or 0

    for row in table.values():
        row["corrections"] = round((row["total_amount"] or 0) - (row["amount"] or 0), 2)

    rows = [table[key] for key in order]
    rows.sort(key=lambda r: (r["employee_name"] or "", r["store"] or "", r["service"] or ""))
    return days, rows

def _number(value):
    return value or 0


def _build_payroll_run_workbook(session, run, items):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"

    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:C2")
    sheet.merge_cells("D1:D2")
    sheet.merge_cells("E1:G1")
    sheet.merge_cells("H1:H2")
    sheet.merge_cells("I1:I2")
    sheet.merge_cells("J1:J2")

    sheet["A1"] = "№ п/п"
    sheet["B1"] = "ФИО"
    sheet["C1"] = "Сумма на карту"
    sheet["D1"] = "ИНН сотрудника"
    sheet["E1"] = "Реквизиты карты"
    sheet["E2"] = "Счёт №"
    sheet["F2"] = "БИК"
    sheet["G2"] = "Банк получателя"
    sheet["H1"] = "Страна гражданства"
    sheet["I1"] = "Получатель 3-лицо"
    sheet["J1"] = "Вид занятости"

    totals_by_employee = {}
    for item in items:
        employee_key = normalize_text(item.employee_name)
        if not employee_key:
            continue
        if employee_key not in totals_by_employee:
            totals_by_employee[employee_key] = {
                "employee_name": item.employee_name,
                "amount": 0,
            }
        totals_by_employee[employee_key]["amount"] += _number(item.total_amount)

    employee_names = sorted(totals_by_employee)
    requisites = {
        normalize_text(row.employee_name): row
        for row in session.query(Requisite).filter(Requisite.is_active == True).all()
    }
    users = {
        normalize_text(row.employee_name): row
        for row in session.query(User).all()
    }

    row_index = 3
    for index, employee_key in enumerate(employee_names, start=1):
        row_data = totals_by_employee[employee_key]
        requisite = requisites.get(employee_key)
        user = users.get(employee_key)

        sheet.cell(row_index, 1, index)
        sheet.cell(row_index, 2, row_data["employee_name"])
        sheet.cell(row_index, 3, row_data["amount"])
        sheet.cell(row_index, 4, str(requisite.inn) if requisite and requisite.inn else None)
        sheet.cell(row_index, 5, str(requisite.account_number) if requisite and requisite.account_number else None)
        sheet.cell(row_index, 6, str(requisite.bik) if requisite and requisite.bik else None)
        sheet.cell(row_index, 7, requisite.bank_name if requisite else None)
        sheet.cell(row_index, 8, citizenship_display(session, user) or None)
        sheet.cell(row_index, 9, requisite.recipient_name if requisite and requisite.recipient_name else None)
        sheet.cell(row_index, 10, None)
        row_index += 1

    total_row = row_index
    sheet.cell(total_row, 2, "ИТОГО")
    if total_row > 3:
        sheet.cell(total_row, 3, f"=SUM(C3:C{total_row - 1})")
    else:
        sheet.cell(total_row, 3, 0)

    for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        for cell in sheet[column]:
            cell.font = Font(name="Calibri", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
        for cell in row:
            cell.font = Font(name="Calibri", size=12, bold=True)

    for cell in sheet[total_row]:
        cell.font = Font(name="Calibri", size=12, bold=True)

    for row in sheet.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    for row in sheet.iter_rows(min_row=3, max_row=total_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "#,##0"

    for column in ["D", "E", "F"]:
        for cell in sheet[column]:
            cell.number_format = "@"

    sheet.column_dimensions["A"].width = 8
    sheet.column_dimensions["B"].width = 44
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 24
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 24
    sheet.column_dimensions["H"].width = 20
    sheet.column_dimensions["I"].width = 44
    sheet.column_dimensions["J"].width = 16

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _export_response(session, run, items):
    output = _build_payroll_run_workbook(session, run, items)
    filename = _export_filename(run)
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _export_filename(run):
    month_names = {
        1: "января",
        2: "февраля",
        3: "марта",
        4: "апреля",
        5: "мая",
        6: "июня",
        7: "июля",
        8: "августа",
        9: "сентября",
        10: "октября",
        11: "ноября",
        12: "декабря",
    }

    if run.date_from.month == run.date_to.month and run.date_from.year == run.date_to.year:
        period = (
            f"с {run.date_from.day:02d} по {run.date_to.day:02d} "
            f"{month_names[run.date_to.month]}"
        )
    else:
        period = (
            f"с {run.date_from.day:02d} {month_names[run.date_from.month]} "
            f"по {run.date_to.day:02d} {month_names[run.date_to.month]}"
        )

    legal_entity = normalize_text(run.legal_entity) or "Без юрлица"
    filename = f"Реестр Лента {period} Оформленные {legal_entity}.xlsx"

    forbidden_chars = '<>:"/\\|?*'
    for char in forbidden_chars:
        filename = filename.replace(char, " ")
    return " ".join(filename.split())


def _user_name(session, user_id):
    if not user_id:
        return None
    user = session.query(User).filter(User.id == user_id).first()
    return user.employee_name if user else None


@router.get("/economist/payroll/runs", response_class=HTMLResponse)
def economist_payroll_runs(
    request: Request,
    message: str = "",
    error: str = "",
    session: Session = Depends(get_db),
    user=Depends(require_payroll_view),
):
    allowed_cities = None if user.is_admin else get_economist_cities(user)

    return templates.TemplateResponse(
        request,
        "payroll_runs.html",
        {
            "run_rows": build_run_rows(session, allowed_cities),
            "error": error or None,
            "message": message or None
        }
    )



@router.get("/economist/payroll/runs/{run_id}", response_class=HTMLResponse)
def economist_payroll_run_detail(
    request: Request,
    run_id: int,
    session: Session = Depends(get_db),
    user=Depends(require_payroll_view),
):
    allowed_cities = None if user.is_admin else get_economist_cities(user)
    run, items = get_run_items(session, run_id, allowed_cities)

    if not run or not items:
        return RedirectResponse(url="/economist/payroll/runs", status_code=302)

    total_hours = sum(item.hours or 0 for item in items)
    total_amount = sum(item.total_amount or 0 for item in items)
    days, rows = build_run_day_rows(run, items)

    return templates.TemplateResponse(
        request,
        "payroll_run_detail.html",
        {
            "user": user,
            "is_economist_view": True,
            "run": run,
            "items": items,
            "days": days,
            "rows": rows,
            "total_hours": total_hours,
            "total_amount": total_amount,
            "closed_by_name": _user_name(session, run.closed_by),
        }
    )




@router.get("/hr/payroll/runs", response_class=HTMLResponse)
def hr_payroll_runs(request: Request, message: str = "", error: str = "", session: Session = Depends(get_db), user=Depends(require_payroll_view)):
    if canonical_role(user) not in {"hr_lead", "hr_manager"}:
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse(request, "payroll_runs.html", {"run_rows": build_run_rows_for_user(session, user), "error": error or None, "message": message or None})


@router.get("/hr/payroll/runs/{run_id}", response_class=HTMLResponse)
def hr_payroll_run_detail(request: Request, run_id: int, session: Session = Depends(get_db), user=Depends(require_payroll_view)):
    if canonical_role(user) not in {"hr_lead", "hr_manager"}:
        return RedirectResponse("/", status_code=302)
    run, items = get_run_items_for_user(session, run_id, user)
    if not run or not items:
        return RedirectResponse("/hr/payroll/runs", status_code=302)
    days, rows = build_run_day_rows(run, items)
    return templates.TemplateResponse(request, "payroll_run_detail.html", {"user": user, "is_economist_view": True, "run": run, "items": items, "days": days, "rows": rows, "total_hours": sum(item.hours or 0 for item in items), "total_amount": sum(item.total_amount or 0 for item in items), "closed_by_name": _user_name(session, run.closed_by)})


@router.get("/hr/payroll/runs/{run_id}/export")
def hr_export_payroll_run(request: Request, run_id: int, session: Session = Depends(get_db), user=Depends(require_payroll_export)):
    if canonical_role(user) not in {"hr_lead", "hr_manager"}:
        return RedirectResponse("/", status_code=302)
    run, items = get_run_items_for_user(session, run_id, user)
    if not run or not items:
        return RedirectResponse("/hr/payroll/runs?error=Нет строк в доступной области", status_code=302)
    return _export_response(session, run, items)

@router.get("/admin/payroll/runs", response_class=HTMLResponse)
def payroll_runs_page(
    request: Request,
    message: str = "",
    error: str = "",
    session: Session = Depends(get_db),
    user=Depends(_payroll_manage),
):
    return templates.TemplateResponse(
        request,
        "payroll_runs.html",
        {
            "run_rows": build_run_rows(session),
            "error": error or None,
            "message": message or None
        }
    )



@router.get("/admin/payroll/runs/{run_id}", response_class=HTMLResponse)
def payroll_run_detail(
    request: Request,
    run_id: int,
    session: Session = Depends(get_db),
    user=Depends(_payroll_manage),
):
    run, items = get_run_items(session, run_id)

    if not run:
        return RedirectResponse(url="/admin/payroll/runs", status_code=302)

    total_hours = sum(item.hours or 0 for item in items)
    total_amount = sum(item.total_amount or 0 for item in items)
    days, rows = build_run_day_rows(run, items)

    return templates.TemplateResponse(
        request,
        "payroll_run_detail.html",
        {
            "user": user,
            "is_economist_view": False,
            "run": run,
            "items": items,
            "days": days,
            "rows": rows,
            "total_hours": total_hours,
            "total_amount": total_amount,
            "closed_by_name": _user_name(session, run.closed_by),
        }
    )



@router.get("/admin/payroll/runs/{run_id}/export")
def export_payroll_run(
    request: Request,
    run_id: int,
    session: Session = Depends(get_db),
    user=Depends(_payroll_manage),
):
    run, items = get_run_items(session, run_id)
    if not run:
        return RedirectResponse(
            url="/admin/payroll/runs?error=Табель не найден",
            status_code=302,
        )
    if not items:
        return RedirectResponse(
            url="/admin/payroll/runs?error=В табеле нет строк для экспорта",
            status_code=302,
        )

    return _export_response(session, run, items)



@router.get("/economist/payroll/runs/{run_id}/export")
def economist_export_payroll_run(
    request: Request,
    run_id: int,
    session: Session = Depends(get_db),
    user=Depends(_payroll_manage),
):
    allowed_cities = None if user.is_admin else get_economist_cities(user)
    run, items = get_run_items(session, run_id, allowed_cities)
    if not run:
        return RedirectResponse(
            url="/economist/payroll/runs?error=Табель не найден",
            status_code=302,
        )
    if not items:
        return RedirectResponse(
            url="/economist/payroll/runs?error=Нет строк по доступным городам",
            status_code=302,
        )

    return _export_response(session, run, items)



@router.post("/admin/payroll/runs/delete")
def delete_payroll_run(
    request: Request,
    run_id: int = Form(...),
    session: Session = Depends(get_db),
    admin=Depends(require_payroll_delete),
):
    run = session.query(PayrollRun).filter(
        PayrollRun.id == run_id
    ).with_for_update().first()

    if not run:
        return RedirectResponse(url="/admin/payroll/runs", status_code=302)

    if run.status != "fixed":
        return RedirectResponse(url="/admin/payroll/runs", status_code=302)

    old_value = {
        "status": run.status,
        "date_from": run.date_from,
        "date_to": run.date_to,
        "legal_entity": run.legal_entity,
    }
    session.query(PayrollRunItem).filter(
        PayrollRunItem.run_id == run.id
    ).delete()

    create_audit_log(
        session,
        request,
        admin,
        "payroll_run_deleted",
        "payroll_run",
        run.id,
        f"{run.date_from} - {run.date_to}",
        old_value=old_value,
    )
    session.delete(run)
    session.commit()

    return RedirectResponse(url="/admin/payroll/runs", status_code=302)



@router.post("/admin/payroll/runs/send")
def send_payroll_run(
    request: Request,
    run_id: int = Form(...),
    session: Session = Depends(get_db),
    user=Depends(current_user),
):
    return_url = "/admin/payroll/runs" if is_superadmin(user) else "/economist/payroll/runs"

    if not has_permission(user, "payroll.send"):
        return RedirectResponse(url="/", status_code=302)

    run = session.query(PayrollRun).filter(
        PayrollRun.id == run_id
    ).with_for_update().first()

    if not run:
        return RedirectResponse(url=return_url, status_code=302)

    if not is_superadmin(user):
        items = session.query(PayrollRunItem).filter(PayrollRunItem.run_id == run.id).all()
        allowed_cities = get_economist_cities(user)
        if not _all_items_accessible_to_cities(session, items, allowed_cities):
            return RedirectResponse(
                url="/economist/payroll/runs?error=Табель содержит строки вне доступных городов",
                status_code=302,
            )

    if run.status == "fixed":
        old_value = {"status": run.status, "sent_at": run.sent_at}
        run.status = "sent"
        run.sent_by = user.id
        run.sent_at = now_utc()
        create_audit_log(
            session,
            request,
            user,
            "payroll_run_sent",
            "payroll_run",
            run.id,
            f"{run.date_from} - {run.date_to}",
            old_value=old_value,
            new_value={"status": run.status, "sent_at": run.sent_at},
        )
        session.commit()

    return RedirectResponse(url=return_url, status_code=302)


@router.post("/admin/payroll/runs/close")
def close_payroll_run_route(
    request: Request,
    run_id: int = Form(...),
    session: Session = Depends(get_db),
    admin=Depends(require_payroll_close),
):
    run = session.query(PayrollRun).filter(
        PayrollRun.id == run_id
    ).with_for_update().first()

    result = close_payroll_run_financials(session, run, admin, request=request)
    if result["closed"]:
        create_audit_log(
            session,
            request,
            admin,
            "payroll_run_closed",
            "payroll_run",
            run.id if run else run_id,
            f"{run.date_from} - {run.date_to}" if run else None,
            old_value={"status": "sent"},
            new_value={
                "status": "closed",
                "closed_at": run.closed_at if run else None,
                "auto_adjustments_applied": result["auto_adjustments_applied"],
                "manual_adjustments_applied": result["manual_adjustments_applied"],
                "lmk_payments_created": result["lmk_payments_created"],
                "lmk_total_amount": result["lmk_total_amount"],
            },
        )
        session.commit()
        message = (
            "Период закрыт. "
            f"Авто: {result['auto_adjustments_applied']}, "
            f"ручные: {result['manual_adjustments_applied']}, "
            f"ЛМК: {result['lmk_payments_created']} "
            f"({result['lmk_total_amount']:.2f})."
        )
        return RedirectResponse(
            url=f"/admin/payroll/runs?message={quote(message)}",
            status_code=302,
        )

    return RedirectResponse(url="/admin/payroll/runs", status_code=302)




