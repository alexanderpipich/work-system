import calendar
from collections import defaultdict
from datetime import date, datetime, timedelta
import os
from pathlib import Path
from urllib.parse import quote, urlencode

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func
from sqlalchemy.orm import Session

from access import apply_shift_scope, get_user_cities, get_user_stores, scope_allows_store
from audit_helpers import create_audit_log
from dependencies import RedirectException, current_user, get_db
from models import (
    EmployeeMonitoringRecommendation,
    EmployeeMonitoringSettings,
    EmployeeStoreAssignment,
    PlanningForm,
    PlanningScenario,
    Shift,
    StoreReconciliation,
    StoreReconciliationSettings,
    User,
)
from rbac import canonical_role, has_permission, is_superadmin, require_permission
from time_helpers import business_today, now_utc
from utils import normalize_text


router = APIRouter()
templates = Jinja2Templates(directory="templates")
_admin_settings = require_permission("admin.settings")
GENERATED_FILES_ROOT = Path(os.getenv("GENERATED_FILES_ROOT", "generated_files"))
PLANNING_DIR = GENERATED_FILES_ROOT / "planning_forms"
RECONCILIATION_DIR = GENERATED_FILES_ROOT / "reconciliations"


def _is_admin(user):
    return is_superadmin(user)


def require_employee_manager(request: Request, user=Depends(current_user)):
    if not has_permission(user, "employees.view"):
        raise RedirectException("/")
    if request.url.path.startswith("/admin") and not _is_admin(user):
        raise RedirectException("/")
    if request.url.path.startswith("/hr") and canonical_role(user) not in {"hr_lead", "hr_manager"}:
        raise RedirectException("/")
    return user


def _context(request, user, section="dashboard", **extra):
    economist = request.url.path.startswith("/economist")
    hr_view = request.url.path.startswith("/hr")
    context = {
        "user": user,
        "section": section,
        "base_path": "/hr/employees" if hr_view else ("/economist/employees" if economist else "/admin/employees"),
        "back_url": "/hr" if hr_view else ("/economist" if economist else "/admin"),
        "is_admin": _is_admin(user),
    }
    context.update(extra)
    return context


def _scope(query, model, user):
    return apply_shift_scope(query, query.session, user, model)


def _scope_allowed(session, user, city, store):
    return scope_allows_store(session, user, city, store)


def _redirect(request, suffix="", **params):
    root = "/hr/employees" if request.url.path.startswith("/hr") else ("/economist/employees" if request.url.path.startswith("/economist") else "/admin/employees")
    url = f"{root}{suffix}"
    clean = {key: value for key, value in params.items() if value not in (None, "")}
    return RedirectResponse(f"{url}?{urlencode(clean)}" if clean else url, status_code=302)


def _parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def reconciliation_ready_date(period_to):
    current = period_to
    added = 0
    while added < 3:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


def is_fixed_reconciliation_period(date_from, date_to):
    last_day = calendar.monthrange(date_from.year, date_from.month)[1]
    return (
        date_from.year == date_to.year
        and date_from.month == date_to.month
        and ((date_from.day == 1 and date_to.day == 15) or (date_from.day == 16 and date_to.day == last_day))
    )


def _stores(session, user):
    query = session.query(Shift.store, Shift.city).filter(Shift.store != None).distinct()
    query = _scope(query, Shift, user)
    return sorted({(row.store, row.city or "") for row in query.all()})


def _users(session, user):
    employee_names = apply_shift_scope(session.query(Shift.employee).distinct(), session, user, Shift)
    return session.query(User).filter(User.employee_name.in_(employee_names)).order_by(User.employee_name.asc()).all()


def _store_city(session, store):
    row = session.query(Shift.city).filter(Shift.store == normalize_text(store), Shift.city != None).first()
    return normalize_text(row[0]) if row else ""


def _scenario_name(session, scenario_id):
    row = session.query(PlanningScenario).filter(PlanningScenario.id == scenario_id).first()
    return row.name if row else "Стандартный"


def _filename_part(value, fallback):
    encoded = quote(normalize_text(value), safe="")
    return encoded[:120] or fallback


def _planning_tree(forms):
    tree = defaultdict(lambda: defaultdict(list))
    for form in forms:
        tree[form.city or "Город не указан"][form.store].append(form)
    return {city: dict(stores) for city, stores in tree.items()}


def _build_planning_file(form, scenario_name):
    PLANNING_DIR.mkdir(parents=True, exist_ok=True)
    safe_employee = _filename_part(form.employee_name, "employee")
    safe_store = _filename_part(form.store, "store")
    filename = f"planning_form_{form.id}_{safe_employee}_{safe_store}_{form.date_from}_{form.date_to}.xlsx"
    path = PLANNING_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Бланк планирования"
    details = [
        ("Сотрудник", form.employee_name), ("Магазин", form.store), ("Город", form.city),
        ("Период", f"{form.date_from:%d.%m.%Y} - {form.date_to:%d.%m.%Y}"), ("Сценарий", scenario_name),
    ]
    for row, values in enumerate(details, start=1):
        ws.cell(row, 1, values[0]).font = Font(bold=True)
        ws.cell(row, 2, values[1])
    ws.append([])
    ws.append(["Дата", "День недели", "Планируемые часы", "Комментарий"])
    for cell in ws[7]:
        cell.font = Font(bold=True)
    current = form.date_from
    while current <= form.date_to:
        ws.append([current, current.strftime("%A"), None, None])
        current += timedelta(days=1)
    ws.freeze_panes = "A8"
    for column, width in {"A": 16, "B": 20, "C": 22, "D": 42}.items():
        ws.column_dimensions[column].width = width
    wb.save(path)
    form.file_path = str(path)
    form.file_name = filename


def _build_reconciliation_file(reconciliation, rows):
    RECONCILIATION_DIR.mkdir(parents=True, exist_ok=True)
    safe_store = _filename_part(reconciliation.store, "store")
    filename = f"reconciliation_{safe_store}_{reconciliation.period_from}_{reconciliation.period_to}.xlsx"
    path = RECONCILIATION_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Сверка"
    ws.append(["Магазин", reconciliation.store])
    ws.append(["Период", f"{reconciliation.period_from:%d.%m.%Y} - {reconciliation.period_to:%d.%m.%Y}"])
    ws.append([])
    ws.append(["ФИО", "Услуга", "Часы"])
    for cell in ws[4]:
        cell.font = Font(bold=True)
    service_totals = defaultdict(float)
    total = 0.0
    for employee, service, hours in rows:
        value = float(hours or 0)
        ws.append([employee, service, value])
        service_totals[service] += value
        total += value
    ws.append([])
    ws.append(["Итоги по услугам"])
    for service, hours in sorted(service_totals.items()):
        ws.append([service, None, hours])
    ws.append(["Общий итог", None, total])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16
    wb.save(path)
    reconciliation.file_path = str(path)


def _monitoring_settings(session):
    settings = session.query(EmployeeMonitoringSettings).first()
    if not settings:
        settings = EmployeeMonitoringSettings(inactive_days=14, minimum_shifts=3, analysis_days=14)
        session.add(settings)
        session.flush()
    return settings


def _create_recommendation(session, request, user, employee, store, city, kind, text_value):
    exists = session.query(EmployeeMonitoringRecommendation).filter(
        EmployeeMonitoringRecommendation.employee_name == employee,
        EmployeeMonitoringRecommendation.store == store,
        EmployeeMonitoringRecommendation.recommendation_type == kind,
        EmployeeMonitoringRecommendation.status == "new",
    ).first()
    if exists:
        return 0
    row = EmployeeMonitoringRecommendation(
        employee_name=employee, store=store, city=city, recommendation_type=kind,
        remove_from_planning=kind == "remove_from_planning", add_to_planning=kind == "add_to_planning",
        recommendation_text=text_value, status="new", created_at=now_utc(),
    )
    session.add(row)
    session.flush()
    create_audit_log(session, request, user, "monitoring_recommendation_created", "employee_monitoring_recommendation", row.id, f"{employee} / {store}")
    return 1


@router.get("/admin/employees", response_class=HTMLResponse)
@router.get("/economist/employees", response_class=HTMLResponse)
@router.get("/hr/employees", response_class=HTMLResponse)
def employees_dashboard(request: Request, user=Depends(require_employee_manager)):
    return templates.TemplateResponse(request, "employees_dashboard.html", _context(request, user))


@router.get("/admin/employees/planning", response_class=HTMLResponse)
@router.get("/economist/employees/planning", response_class=HTMLResponse)
@router.get("/hr/employees/planning", response_class=HTMLResponse)
def planning_page(request: Request, view: str = "list", message: str = "", session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    forms = _scope(session.query(PlanningForm), PlanningForm, user).order_by(PlanningForm.id.desc()).all()
    assignments = _scope(session.query(EmployeeStoreAssignment), EmployeeStoreAssignment, user).order_by(EmployeeStoreAssignment.city, EmployeeStoreAssignment.store, EmployeeStoreAssignment.employee_name).all()
    scenarios = session.query(PlanningScenario).order_by(PlanningScenario.is_active.desc(), PlanningScenario.name).all()
    return templates.TemplateResponse(request, "employees_planning.html", _context(request, user, "planning", forms=forms, assignments=assignments, scenarios=scenarios, users=_users(session, user), stores=_stores(session, user), tree=_planning_tree(forms), view=view, message=message))


@router.post("/admin/employees/planning/assignments/add")
@router.post("/economist/employees/planning/assignments/add")
@router.post("/hr/employees/planning/assignments/add")
def assignment_add(request: Request, employee_name: str = Form(...), store: str = Form(...), comment: str = Form(default=""), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    city = _store_city(session, store)
    if not _scope_allowed(session, user, city, store):
        return _redirect(request, "/planning", message="Нет доступа к городу")
    existing = session.query(EmployeeStoreAssignment).filter(EmployeeStoreAssignment.employee_name == normalize_text(employee_name), EmployeeStoreAssignment.store == normalize_text(store)).first()
    if existing:
        existing.is_active = True
        existing.comment = normalize_text(comment) or None
        existing.city = city
        action = "planning_assignment_updated"
        row = existing
    else:
        person = session.query(User).filter(User.employee_name == normalize_text(employee_name)).first()
        row = EmployeeStoreAssignment(user_id=person.id if person else None, employee_name=normalize_text(employee_name), store=normalize_text(store), city=city, is_active=True, created_by=user.id, created_at=now_utc(), comment=normalize_text(comment) or None)
        session.add(row); session.flush(); action = "planning_assignment_created"
    create_audit_log(session, request, user, action, "employee_store_assignment", row.id, f"{row.employee_name} / {row.store}")
    session.commit()
    return _redirect(request, "/planning", message="Назначение сохранено")


@router.post("/admin/employees/planning/assignments/update")
@router.post("/economist/employees/planning/assignments/update")
@router.post("/hr/employees/planning/assignments/update")
def assignment_update(request: Request, assignment_id: int = Form(...), is_active: str = Form(default=""), comment: str = Form(default=""), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    row = session.query(EmployeeStoreAssignment).filter(EmployeeStoreAssignment.id == assignment_id).first()
    if not row or not _scope_allowed(session, user, row.city, row.store): return _redirect(request, "/planning", message="Назначение недоступно")
    row.is_active = is_active == "1"; row.comment = normalize_text(comment) or None
    create_audit_log(session, request, user, "planning_assignment_updated", "employee_store_assignment", row.id, f"{row.employee_name} / {row.store}")
    session.commit(); return _redirect(request, "/planning", message="Назначение обновлено")


@router.post("/admin/employees/planning/scenarios/add")
def scenario_add(request: Request, name: str = Form(...), description: str = Form(default=""), template_type: str = Form(default="standard"), email_subject_template: str = Form(default=""), email_body_template: str = Form(default=""), session: Session = Depends(get_db), user=Depends(_admin_settings)):
    row = PlanningScenario(name=normalize_text(name), description=normalize_text(description) or None, template_type=normalize_text(template_type) or "standard", email_subject_template=normalize_text(email_subject_template) or None, email_body_template=normalize_text(email_body_template) or None, is_active=True, created_at=now_utc())
    session.add(row); session.flush(); create_audit_log(session, request, user, "planning_scenario_created", "planning_scenario", row.id, row.name); session.commit()
    return _redirect(request, "/planning", message="Сценарий создан")


@router.post("/admin/employees/planning/assignments/delete")
@router.post("/economist/employees/planning/assignments/delete")
@router.post("/hr/employees/planning/assignments/delete")
def assignment_delete(request: Request, assignment_id: int = Form(...), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    row = session.query(EmployeeStoreAssignment).filter(EmployeeStoreAssignment.id == assignment_id).first()
    if row and _scope_allowed(session, user, row.city, row.store):
        row.is_active = False
        create_audit_log(session, request, user, "planning_assignment_updated", "employee_store_assignment", row.id, f"{row.employee_name} / {row.store}")
        session.commit()
    return _redirect(request, "/planning", message="Назначение деактивировано")


@router.post("/admin/employees/planning/scenarios/update")
def scenario_update(request: Request, scenario_id: int = Form(...), name: str = Form(...), description: str = Form(default=""), template_type: str = Form(default="standard"), is_active: str = Form(default=""), session: Session = Depends(get_db), user=Depends(_admin_settings)):
    row = session.query(PlanningScenario).filter(PlanningScenario.id == scenario_id).first()
    if row:
        row.name = normalize_text(name); row.description = normalize_text(description) or None; row.template_type = normalize_text(template_type) or "standard"; row.is_active = is_active == "1"; row.updated_at = now_utc()
        create_audit_log(session, request, user, "planning_scenario_updated", "planning_scenario", row.id, row.name); session.commit()
    return _redirect(request, "/planning", message="Сценарий обновлён")


@router.post("/admin/employees/planning/scenarios/delete")
def scenario_delete(request: Request, scenario_id: int = Form(...), session: Session = Depends(get_db), user=Depends(_admin_settings)):
    row = session.query(PlanningScenario).filter(PlanningScenario.id == scenario_id).first()
    if row:
        row.is_active = False; row.updated_at = now_utc(); create_audit_log(session, request, user, "planning_scenario_updated", "planning_scenario", row.id, row.name); session.commit()
    return _redirect(request, "/planning", message="Сценарий деактивирован")

@router.post("/admin/employees/planning/forms/create")
@router.post("/economist/employees/planning/forms/create")
@router.post("/hr/employees/planning/forms/create")
def planning_form_create(request: Request, scenario_id: int = Form(default=0), employee_name: str = Form(...), store: str = Form(...), date_from: str = Form(...), date_to: str = Form(...), email_to: str = Form(default=""), comment: str = Form(default=""), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    city = _store_city(session, store)
    if not _scope_allowed(session, user, city, store): return _redirect(request, "/planning", message="Нет доступа к городу")
    start, end = _parse_date(date_from), _parse_date(date_to)
    if start > end:
        return _redirect(request, "/planning", message="Дата начала позже даты окончания")
    person = session.query(User).filter(User.employee_name == normalize_text(employee_name)).first()
    scenario = session.query(PlanningScenario).filter(PlanningScenario.id == scenario_id, PlanningScenario.is_active == True).first() if scenario_id else None
    row = PlanningForm(scenario_id=scenario.id if scenario else None, employee_name=normalize_text(employee_name), user_id=person.id if person else None, store=normalize_text(store), city=city, date_from=start, date_to=end, status="created", created_by=user.id, created_at=now_utc(), email_to=normalize_text(email_to) or None, email_subject=(scenario.email_subject_template if scenario else None), comment=normalize_text(comment) or None)
    session.add(row); session.flush(); _build_planning_file(row, scenario.name if scenario else "Стандартный")
    create_audit_log(session, request, user, "planning_form_created", "planning_form", row.id, f"{row.employee_name} / {row.store}"); session.commit()
    return _redirect(request, "/planning", message="Бланк сформирован")


def _get_planning_form(session, user, form_id):
    row = session.query(PlanningForm).filter(PlanningForm.id == form_id).first()
    return row if row and _scope_allowed(session, user, row.city, row.store) else None


@router.get("/admin/employees/planning/forms/{form_id}/download")
@router.get("/economist/employees/planning/forms/{form_id}/download")
@router.get("/hr/employees/planning/forms/{form_id}/download")
def planning_form_download(form_id: int, session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    row = _get_planning_form(session, user, form_id)
    if not row or not row.file_path or not Path(row.file_path).is_file(): return RedirectResponse("/", status_code=302)
    return FileResponse(row.file_path, filename=row.file_name)


@router.post("/admin/employees/planning/forms/{form_id}/{action}")
@router.post("/economist/employees/planning/forms/{form_id}/{action}")
@router.post("/hr/employees/planning/forms/{form_id}/{action}")
def planning_form_action(request: Request, form_id: int, action: str, session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    row = _get_planning_form(session, user, form_id)
    if not row or action not in {"send", "cancel"}: return _redirect(request, "/planning")
    row.status = "sent" if action == "send" else "cancelled"
    if action == "send": row.sent_by = user.id; row.sent_at = now_utc()
    create_audit_log(session, request, user, f"planning_form_{'sent' if action == 'send' else 'cancelled'}", "planning_form", row.id, f"{row.employee_name} / {row.store}"); session.commit()
    return _redirect(request, "/planning", message="Статус бланка обновлён")


@router.get("/admin/employees/monitoring", response_class=HTMLResponse)
@router.get("/economist/employees/monitoring", response_class=HTMLResponse)
@router.get("/hr/employees/monitoring", response_class=HTMLResponse)
def monitoring_page(request: Request, tab: str = "new", message: str = "", session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    query = _scope(session.query(EmployeeMonitoringRecommendation), EmployeeMonitoringRecommendation, user)
    if tab == "inactive": query = query.filter(EmployeeMonitoringRecommendation.recommendation_type == "remove_from_planning")
    elif tab == "missing": query = query.filter(EmployeeMonitoringRecommendation.recommendation_type == "add_to_planning")
    elif tab != "all": query = query.filter(EmployeeMonitoringRecommendation.status == "new")
    rows = query.order_by(EmployeeMonitoringRecommendation.id.desc()).all()
    return templates.TemplateResponse(request, "employees_monitoring.html", _context(request, user, "monitoring", rows=rows, tab=tab, settings=_monitoring_settings(session), message=message))


@router.get("/admin/employees/monitoring/settings", response_class=HTMLResponse)
def monitoring_settings_page(request: Request, session: Session = Depends(get_db), user=Depends(_admin_settings)):
    return templates.TemplateResponse(request, "employees_monitoring.html", _context(request, user, "monitoring", rows=[], tab="settings", settings=_monitoring_settings(session), message=""))

@router.post("/admin/employees/monitoring/settings")
def monitoring_settings_save(request: Request, inactive_days: int = Form(...), minimum_shifts: int = Form(...), analysis_days: int = Form(...), session: Session = Depends(get_db), user=Depends(_admin_settings)):
    row = _monitoring_settings(session); row.inactive_days=max(inactive_days,1); row.minimum_shifts=max(minimum_shifts,1); row.analysis_days=max(analysis_days,1); row.updated_by=user.id; row.updated_at=now_utc(); session.commit()
    return _redirect(request, "/monitoring", message="Настройки сохранены")


@router.post("/admin/employees/monitoring/generate")
@router.post("/economist/employees/monitoring/generate")
@router.post("/hr/employees/monitoring/generate")
def monitoring_generate(request: Request, session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    settings = _monitoring_settings(session); today = business_today(); created = 0
    assignments = _scope(session.query(EmployeeStoreAssignment).filter(EmployeeStoreAssignment.is_active == True), EmployeeStoreAssignment, user).all()
    active_keys = {(a.employee_name, a.store) for a in assignments}
    for item in assignments:
        last_date = session.query(func.max(Shift.shift_date)).filter(Shift.employee == item.employee_name, Shift.store == item.store).scalar()
        if not last_date or last_date <= today - timedelta(days=settings.inactive_days):
            created += _create_recommendation(session, request, user, item.employee_name, item.store, item.city, "remove_from_planning", f"Сотрудник закреплён за магазином, но не имеет смен {settings.inactive_days} дней. Возможно его следует убрать из бланка планирования.")
    query = session.query(Shift.employee, Shift.store, Shift.city, func.count(Shift.id).label("count")).filter(Shift.shift_date >= today - timedelta(days=settings.analysis_days - 1)).group_by(Shift.employee, Shift.store, Shift.city)
    query = _scope(query, Shift, user)
    for item in query.having(func.count(Shift.id) >= settings.minimum_shifts).all():
        if (item.employee, item.store) not in active_keys:
            created += _create_recommendation(session, request, user, item.employee, item.store, item.city, "add_to_planning", f"Сотрудник отработал {item.count} смен за последние {settings.analysis_days} дней, но отсутствует в планировании. Возможно его следует добавить.")
    session.commit(); return _redirect(request, "/monitoring", message=f"Создано рекомендаций: {created}")


@router.post("/admin/employees/monitoring/process")
@router.post("/economist/employees/monitoring/process")
@router.post("/hr/employees/monitoring/process")
def monitoring_process(request: Request, recommendation_id: int = Form(...), status: str = Form(...), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    row = session.query(EmployeeMonitoringRecommendation).filter(EmployeeMonitoringRecommendation.id == recommendation_id).first()
    if row and _scope_allowed(session, user, row.city, row.store) and status in {"accepted", "dismissed"}: row.status=status; row.processed_by=user.id; row.processed_at=now_utc(); session.commit()
    return _redirect(request, "/monitoring", message="Рекомендация обработана")


@router.get("/admin/employees/reconciliations", response_class=HTMLResponse)
@router.get("/economist/employees/reconciliations", response_class=HTMLResponse)
@router.get("/hr/employees/reconciliations", response_class=HTMLResponse)
def reconciliations_page(request: Request, message: str = "", session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    rows = _scope(session.query(StoreReconciliation), StoreReconciliation, user).order_by(StoreReconciliation.id.desc()).all()
    settings = _scope(session.query(StoreReconciliationSettings), StoreReconciliationSettings, user).order_by(StoreReconciliationSettings.store).all()
    return templates.TemplateResponse(request, "employees_reconciliations.html", _context(request, user, "reconciliations", rows=rows, settings=settings, stores=_stores(session,user), message=message))


@router.post("/admin/employees/reconciliations/settings")
@router.post("/economist/employees/reconciliations/settings")
@router.post("/hr/employees/reconciliations/settings")
def reconciliation_settings(request: Request, store: str = Form(...), email: str = Form(default=""), enabled: str = Form(default=""), comment: str = Form(default=""), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    city=_store_city(session,store)
    if not _scope_allowed(session,user,city,store): return _redirect(request,"/reconciliations",message="Нет доступа к городу")
    row=session.query(StoreReconciliationSettings).filter(StoreReconciliationSettings.store==normalize_text(store)).first() or StoreReconciliationSettings(store=normalize_text(store)); session.add(row); row.city=city; row.email=normalize_text(email) or None; row.enabled=enabled=="1"; row.comment=normalize_text(comment) or None; session.commit()
    return _redirect(request,"/reconciliations",message="Настройки сохранены")


@router.post("/admin/employees/reconciliations/generate")
@router.post("/economist/employees/reconciliations/generate")
@router.post("/hr/employees/reconciliations/generate")
def reconciliation_generate(request: Request, store: str = Form(...), date_from: str = Form(...), date_to: str = Form(...), session: Session = Depends(get_db), user=Depends(require_employee_manager)):
    city=_store_city(session,store)
    if not _scope_allowed(session,user,city,store): return _redirect(request,"/reconciliations",message="Нет доступа к городу")
    start,end=_parse_date(date_from),_parse_date(date_to)
    if not is_fixed_reconciliation_period(start,end):return _redirect(request,"/reconciliations",message="Допустимы только периоды 01–15 или 16–конец месяца")
    if business_today() < reconciliation_ready_date(end):return _redirect(request,"/reconciliations",message=f"Сверку можно сформировать с {reconciliation_ready_date(end):%d.%m.%Y}")
    existing=session.query(StoreReconciliation).filter(StoreReconciliation.store==normalize_text(store),StoreReconciliation.period_from==start,StoreReconciliation.period_to==end).first()
    if existing:return _redirect(request,"/reconciliations",message="Сверка за период уже существует")
    data=session.query(Shift.employee,Shift.service,func.sum(Shift.hours)).filter(Shift.store==normalize_text(store),Shift.shift_date>=start,Shift.shift_date<=end).group_by(Shift.employee,Shift.service).order_by(Shift.employee,Shift.service).all()
    row=StoreReconciliation(store=normalize_text(store),city=city,period_from=start,period_to=end,status="generated",created_at=now_utc()); session.add(row);session.flush();_build_reconciliation_file(row,data);create_audit_log(session,request,user,"reconciliation_generated","store_reconciliation",row.id,row.store);session.commit()
    return _redirect(request,"/reconciliations",message="Сверка сформирована")


def _reconciliation(session,user,row_id):
    row=session.query(StoreReconciliation).filter(StoreReconciliation.id==row_id).first(); return row if row and _scope_allowed(session,user,row.city,row.store) else None


@router.get("/admin/employees/reconciliations/{row_id}/download")
@router.get("/economist/employees/reconciliations/{row_id}/download")
@router.get("/hr/employees/reconciliations/{row_id}/download")
def reconciliation_download(row_id:int,session:Session=Depends(get_db),user=Depends(require_employee_manager)):
    row=_reconciliation(session,user,row_id)
    if not row or not row.file_path or not Path(row.file_path).is_file():return RedirectResponse("/",status_code=302)
    return FileResponse(row.file_path,filename=Path(row.file_path).name)


@router.post("/admin/employees/reconciliations/{row_id}/{action}")
@router.post("/economist/employees/reconciliations/{row_id}/{action}")
@router.post("/hr/employees/reconciliations/{row_id}/{action}")
def reconciliation_action(request:Request,row_id:int,action:str,session:Session=Depends(get_db),user=Depends(require_employee_manager)):
    row=_reconciliation(session,user,row_id)
    if row and action in {"ready","sent"}: row.status="ready_to_send" if action=="ready" else "sent"; row.sent_at=now_utc() if action=="sent" else None; create_audit_log(session,request,user,"reconciliation_sent" if action=="sent" else "reconciliation_ready","store_reconciliation",row.id,row.store);session.commit()
    return _redirect(request,"/reconciliations",message="Статус сверки обновлён")

