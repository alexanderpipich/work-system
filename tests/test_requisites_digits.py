"""ИНН/счёт/БИК не должны портиться ни на входе, ни на выгрузке в реестр."""

import os
import unittest

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SECRET_KEY", "test-secret")

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import PayrollRunItem, Requisite, User
from routers.payroll_runs import _build_payroll_run_workbook
from routers.requisites import _clean_identifiers
from utils import (ACCOUNT_LENGTH, BIK_LENGTH, INN_LENGTHS, is_valid_requisite,
                   repair_digits, repair_requisite, requisite_issues)

# Реальные значения из разбора порчи реестра: у Ивановой в базе всё целое,
# у тестового Хисрова — уже испорчено ручным вводом.
IVANOVA = ("471605460681", "40817810055192444131", "044030653")
HISROV = ("780000000000.0", "4e+19", "40000013.0")


def _session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


class ValidatorTests(unittest.TestCase):
    """Единый критерий годности — общий с экраном невалидных реквизитов (002)."""

    def test_correct_requisites_pass(self):
        self.assertTrue(is_valid_requisite(*IVANOVA))
        self.assertEqual(requisite_issues(*IVANOVA), [])

    def test_inn_of_legal_entity_is_ten_digits(self):
        self.assertTrue(is_valid_requisite("7710140679", IVANOVA[1], IVANOVA[2]))

    def test_lengths_are_checked(self):
        self.assertIn("Номер счёта: 15 цифр, ожидалось 20",
                      requisite_issues(IVANOVA[0], "408178100551924", IVANOVA[2]))
        self.assertIn("БИК: 8 цифр, ожидалось 9",
                      requisite_issues(IVANOVA[0], IVANOVA[1], "04403065"))

    def test_non_digits_rejected(self):
        self.assertFalse(is_valid_requisite("47160546068X", IVANOVA[1], IVANOVA[2]))

    def test_empty_reported(self):
        self.assertIn("БИК не заполнен", requisite_issues(IVANOVA[0], IVANOVA[1], ""))

    def test_scientific_notation_is_lost_data(self):
        issues = requisite_issues(IVANOVA[0], "4.08e+19", IVANOVA[2])
        self.assertTrue(any("утеряны" in i for i in issues), issues)

    def test_float_tail_is_recoverable_and_accepted(self):
        # "780000000000.0" — хвост от Excel, цифры целы; после нормализации годен.
        self.assertTrue(is_valid_requisite("780000000000.0", IVANOVA[1], IVANOVA[2]))

    def test_surrounding_spaces_tolerated(self):
        self.assertTrue(is_valid_requisite(*(" %s " % v for v in IVANOVA)))


class InputCleaningTests(unittest.TestCase):
    """Вход ручных форм: хвост отрезаем, заведомо утерянное не сохраняем."""

    def test_float_tail_stripped(self):
        values, refusal = _clean_identifiers("780000000000.0", IVANOVA[1], "044030653.0")
        self.assertIsNone(refusal)
        self.assertEqual(values[0], "780000000000")
        self.assertEqual(values[2], "044030653")

    def test_leading_zeros_survive(self):
        values, _ = _clean_identifiers(IVANOVA[0], IVANOVA[1], "044030653")
        self.assertEqual(values[2], "044030653")

    def test_bik_leading_zero_restored_after_float_damage(self):
        """Ввод «44030653.0» обязан сохраниться как «044030653».

        Регрессия 004: на входе стоял голый normalize_digits — он отрезал «.0»,
        но съеденный float'ом ведущий ноль не возвращал, и в базу ложился БИК
        из 8 цифр. БИК часто начинается с нуля, так что это боевые данные.
        """
        values, refusal = _clean_identifiers(IVANOVA[0], IVANOVA[1], "44030653.0")
        self.assertIsNone(refusal)
        self.assertEqual(values[2], "044030653")
        self.assertEqual(len(values[2]), 9)

    def test_inn_leading_zero_restored(self):
        values, _ = _clean_identifiers("71605460681", IVANOVA[1], IVANOVA[2])
        self.assertEqual(values[0], "071605460681")

    def test_account_keeps_all_twenty_digits(self):
        values, _ = _clean_identifiers(IVANOVA[0], IVANOVA[1], IVANOVA[2])
        self.assertEqual(values[1], IVANOVA[1])
        self.assertEqual(len(values[1]), 20)

    def test_short_account_is_never_padded(self):
        """Счёт короче 20 остаётся как есть — ноль дописать значит выдумать номер.

        Такой реквизит обязан уйти на экран невалидных, а не притвориться годным.
        """
        short = "4081781005519244413"
        values, refusal = _clean_identifiers(IVANOVA[0], short, IVANOVA[2])
        self.assertIsNone(refusal)
        self.assertEqual(values[1], short)
        self.assertFalse(values[1].startswith("0"))
        self.assertFalse(is_valid_requisite(*values))

    def test_repaired_values_are_valid(self):
        values, _ = _clean_identifiers("471605460681.0", IVANOVA[1], "44030653.0")
        self.assertTrue(is_valid_requisite(*values), requisite_issues(*values))

    def test_scientific_notation_refused(self):
        _, refusal = _clean_identifiers(IVANOVA[0], "4.08e+19", IVANOVA[2])
        self.assertIsNotNone(refusal)
        self.assertIn("номер счёта", refusal)

    def test_partial_requisite_still_saves(self):
        # Реквизит бывает заполнен частично — блокировать сохранение нельзя.
        _, refusal = _clean_identifiers("", "", "")
        self.assertIsNone(refusal)


class RepairTests(unittest.TestCase):
    """Починка того, что уже испорчено в базе проходом через float."""

    # Ровно то, что лежало в проде у Ивановой (реестр от 13.08.2026).
    PROD_BROKEN = ("471605460681.0", "4.0817810055192445e+19", "44030653.0")

    def test_prod_case_inn_and_bik_repaired_account_not(self):
        inn, account, bik = repair_requisite(*self.PROD_BROKEN)
        self.assertEqual(inn, "471605460681", "хвост .0 не отрезан")
        self.assertEqual(bik, "044030653", "ведущий ноль БИК не восстановлен")
        self.assertIsNone(account, "счёт с утерянными цифрами нельзя объявлять починенным")

    def test_account_is_never_padded_with_zeros(self):
        """Счёт начинается с балансового счёта (40817…), ведущего нуля там не бывает.

        Дописать ноль — выдумать номер, который пройдёт проверку длины и уйдёт
        в банк неверным. Такой реквизит обязан попасть в переввод, а не «починиться».
        """
        self.assertIsNone(repair_digits("4081781005519244413", ACCOUNT_LENGTH,
                                        pad_leading_zeros=False))
        _, account, _ = repair_requisite(IVANOVA[0], "4081781005519244413", IVANOVA[2])
        self.assertIsNone(account)

    def test_inn_and_bik_leading_zeros_restored(self):
        # Длины ИНН (10/12) и БИК (9) известны заранее — ноль восстанавливается однозначно.
        self.assertEqual(repair_digits("71605460681", INN_LENGTHS), "071605460681")
        self.assertEqual(repair_digits("44030653", BIK_LENGTH), "044030653")

    def test_intact_values_untouched(self):
        self.assertEqual(repair_requisite(*IVANOVA), IVANOVA)

    def test_scientific_notation_never_repaired(self):
        self.assertIsNone(repair_digits("4.08e+19", ACCOUNT_LENGTH, pad_leading_zeros=False))
        self.assertIsNone(repair_digits("7.8e+11", INN_LENGTHS))

    def test_garbage_not_repaired(self):
        self.assertIsNone(repair_digits("47160546068X", INN_LENGTHS))
        self.assertIsNone(repair_digits("", INN_LENGTHS))
        self.assertIsNone(repair_digits(None, INN_LENGTHS))

    def test_too_long_left_alone(self):
        # Длиннее максимума — это не потеря нуля, а что-то другое; не наше дело.
        self.assertIsNone(repair_digits("4718123456789012", INN_LENGTHS))

    def test_repair_is_idempotent(self):
        once = repair_requisite(*self.PROD_BROKEN)
        twice = repair_requisite(once[0], once[1] or self.PROD_BROKEN[1], once[2])
        self.assertEqual(once, twice)

    def test_repaired_values_pass_the_validator(self):
        inn, _, bik = repair_requisite(*self.PROD_BROKEN)
        self.assertEqual(requisite_issues(inn, IVANOVA[1], bik), [])


class OpenpyxlRoundTripTests(unittest.TestCase):
    """Строка-из-цифр обязана пережить запись и чтение без искажения.

    Ловит регресс, если openpyxl когда-нибудь вернёт угадывание типов: тогда
    20-значный счёт схлопнется во float и младшие цифры пропадут.
    """

    def test_digit_strings_survive_round_trip(self):
        workbook = Workbook()
        sheet = workbook.active
        values = ["40817810055192444131", "471605460681", "044030653", "00000000000000000001"]
        for index, value in enumerate(values, start=1):
            sheet.cell(index, 1, value)
        for cell in sheet["A"]:
            cell.number_format = "@"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        reread = load_workbook(buffer).active

        for index, value in enumerate(values, start=1):
            with self.subTest(value=value):
                cell = reread.cell(index, 1)
                self.assertEqual(cell.value, value)
                self.assertEqual(cell.data_type, "s", "значение перестало быть текстом")


class RegistryExportTests(unittest.TestCase):
    """Реестр выплат обязан выдавать ровно то, что лежит в базе."""

    def setUp(self):
        self.session = _session()
        for name, (inn, account, bik) in (("Иванова Т.М.", IVANOVA), ("Хисров Х.", HISROV)):
            self.session.add(User(employee_name=name, phone="7999%s" % len(name),
                                  password_hash="x"))
            self.session.add(Requisite(employee_name=name, inn=inn, account_number=account,
                                       bik=bik, bank_name="Банк", is_active=True))
        self.session.commit()

    def _export(self, names=("Иванова Т.М.", "Хисров Х.")):
        class Run:
            id = 1
            legal_entity = "ООО"
            period_start = period_end = None
            status = "draft"

        items = [PayrollRunItem(employee_name=name, total_amount=1000 * (index + 1))
                 for index, name in enumerate(names)]
        return load_workbook(_build_payroll_run_workbook(self.session, Run(), items)).active

    def test_intact_requisites_export_intact(self):
        sheet = self._export()
        self.assertEqual(sheet.cell(3, 2).value, "Иванова Т.М.")
        self.assertEqual(sheet.cell(3, 4).value, IVANOVA[0])
        self.assertEqual(sheet.cell(3, 5).value, IVANOVA[1], "20-значный счёт исказился")
        self.assertEqual(sheet.cell(3, 6).value, IVANOVA[2])

    def test_identifier_cells_stay_text(self):
        sheet = self._export()
        for column in (4, 5, 6):
            with self.subTest(column=column):
                self.assertEqual(sheet.cell(3, column).data_type, "s")

    def test_no_scientific_notation_for_intact_data(self):
        sheet = self._export()
        self.assertNotIn("e+", str(sheet.cell(3, 5).value))
        self.assertFalse(str(sheet.cell(3, 4).value).endswith(".0"))

    def test_float_damaged_input_reaches_registry_whole(self):
        """Сквозной путь порченого ввода: форма → база → реестр, ноль БИК на месте.

        Порознь и ввод, и выгрузка были зелёными, а ноль всё равно терялся —
        ловит только проверка всей цепочки целиком.
        """
        values, refusal = _clean_identifiers("471605460681.0", IVANOVA[1], "44030653.0")
        self.assertIsNone(refusal)
        self.session.add(User(employee_name="Петров П.П.", phone="79990000001",
                              password_hash="x"))
        self.session.add(Requisite(employee_name="Петров П.П.", inn=values[0],
                                   account_number=values[1], bik=values[2],
                                   bank_name="Банк", is_active=True))
        self.session.commit()

        sheet = self._export(names=("Петров П.П.",))
        self.assertEqual(sheet.cell(3, 4).value, "471605460681")
        self.assertEqual(sheet.cell(3, 5).value, IVANOVA[1])
        self.assertEqual(sheet.cell(3, 6).value, "044030653")
        self.assertEqual(sheet.cell(3, 6).data_type, "s")


if __name__ == "__main__":
    unittest.main()
