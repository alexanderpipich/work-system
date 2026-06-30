"""
Генератор XLS-сверки (openpyxl) — повтор образца
`сверка_ТК-007_и_ООО_ПРОГРЕСС_-_апрель_2026_-_202ч.xlsx` 1-в-1.

Только ЧАСЫ, без денег. Один лист, сверху вниз:
  1. Шапка-итог (период | Магазин | Поставщик | СУММА по всем услугам | всего часов).
  2. Свод по услугам (на каждую услугу пара строк заголовок/значение).
  3. Детализация по сотрудникам (блок на сотрудника: заголовок, строки смен,
     строка-итог по сотруднику).
Контроль: Σ свод = Σ итогов сотрудников = ИТОГО в шапке.

Источники данных передаёт вызывающий (helpers) — модуль только верстает XLS.
«Поставщик (Исполнитель)» = юрлицо-константа (как в БП), «Тип заявки» =
Shift.request_type.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Зелёная шапка таблиц — как в образце.
_HEADER_FILL = PatternFill("solid", fgColor="C6E0B4")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def format_hours(h):
    """Часы: целое без дробной части (202), иначе с дробью."""
    value = float(h or 0)
    return int(value) if value.is_integer() else round(value, 2)


def _style_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.fill = _HEADER_FILL
    cell.border = _BORDER
    cell.alignment = _CENTER


def _style_value(cell, *, bold=False, align=None):
    cell.font = Font(name="Calibri", size=11, bold=bold)
    cell.border = _BORDER
    cell.alignment = align or _LEFT


def _aggregate_by_service(shifts):
    """{услуга: сумма часов}, по возрастанию названия услуги."""
    totals = {}
    for s in shifts:
        totals[s.service] = totals.get(s.service, 0.0) + float(s.hours or 0)
    return dict(sorted(totals.items(), key=lambda kv: kv[0]))


def _group_by_employee(shifts):
    """{ФИО: [смены...]}, сотрудники по алфавиту, смены по дате."""
    groups = {}
    for s in shifts:
        groups.setdefault(s.employee, []).append(s)
    ordered = {}
    for name in sorted(groups.keys()):
        ordered[name] = sorted(groups[name], key=lambda s: (s.shift_date, s.service))
    return ordered


def build_reconciliation_xls(*, tk, store_name, legal_entity, period_label, shifts):
    """
    Собрать XLS-сверку. `shifts` — смены ТК за период, БЕЗ смен без плана
    (фильтрация на стороне вызывающего). Возвращает bytes.

    Поля каждой смены: shift_date, employee, service, request_type, hours.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    store_name = store_name or f"ТК-{tk:03d}"
    total_hours = sum(float(s.hours or 0) for s in shifts)

    # 1) Шапка-итог (r1 заголовки, r2 значения).
    head_titles = ["период", "Магазин", "Поставщик (Исполнитель)", "Услуга", "ИТОГО часов"]
    for col, title in enumerate(head_titles, start=1):
        _style_header(ws.cell(1, col, title))
    head_values = [period_label, store_name, legal_entity, "СУММА по всем услугам", format_hours(total_hours)]
    for col, value in enumerate(head_values, start=1):
        _style_value(ws.cell(2, col, value), bold=True,
                     align=_CENTER if col == 5 else _LEFT)

    # 2) Свод по услугам (с r5), мини-блоки через пустую строку.
    by_service = _aggregate_by_service(shifts)
    row = 5
    for service, hours in by_service.items():
        _style_header(ws.cell(row, 4, "Услуга"))
        _style_header(ws.cell(row, 5, "ИТОГО часов"))
        _style_value(ws.cell(row + 1, 4, service))
        _style_value(ws.cell(row + 1, 5, format_hours(hours)), align=_CENTER)
        row += 3  # пара строк + пустая

    # 3) Детализация по сотрудникам.
    row += 1
    detail_titles = ["Дата", "Магазин", "Поставщик (Исполнитель)",
                     "ФИО сотрудника Исполнителя", "Услуга", "Тип заявки", "Часы"]
    for name, emp_shifts in _group_by_employee(shifts).items():
        for col, title in enumerate(detail_titles, start=1):
            _style_header(ws.cell(row, col, title))
        row += 1
        emp_total = 0.0
        for s in emp_shifts:
            emp_total += float(s.hours or 0)
            values = [
                s.shift_date.strftime("%d.%m.%Y"),
                store_name,
                legal_entity,
                s.employee,
                s.service,
                s.request_type,
                format_hours(s.hours),
            ]
            for col, value in enumerate(values, start=1):
                _style_value(ws.cell(row, col, value),
                             align=_CENTER if col in (1, 7) else _LEFT)
            row += 1
        # Строка-итог сотрудника: сумма только в колонке «Часы».
        _style_value(ws.cell(row, 7, format_hours(emp_total)), bold=True, align=_CENTER)
        row += 2  # итог + пустая

    widths = {"A": 14, "B": 40, "C": 26, "D": 36, "E": 36, "F": 18, "G": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_filename(tk, legal_entity, month_label, total_hours):
    """`сверка_ТК-007_и_ООО_ПРОГРЕСС_-_апрель_2026_-_202ч.xlsx`."""
    entity = (legal_entity or "").replace(" ", "_")
    month = (month_label or "").replace(" ", "_")
    hours = format_hours(total_hours)
    return f"сверка_ТК-{tk:03d}_и_{entity}_-_{month}_-_{hours}ч.xlsx"
